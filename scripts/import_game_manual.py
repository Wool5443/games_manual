import re
import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = BASE_DIR / "Game_manual.xlsx"
DATABASE_PATH = BASE_DIR / "instance" / "games.db"
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from app import DEFAULT_AGE_OPTIONS, DEFAULT_GAME_TYPES

DEFAULT_LOCATIONS = {"Не имеет значения", "Помещение", "Улица"}

TEXT_COLUMNS = {
    "C": "title",
    "D": "game_type",
    "E": "goal",
    "F": "participants",
    "G": "age_category",
    "H": "duration",
    "I": "location",
    "J": "equipment",
    "K": "rules",
    "L": "rules_link",
    "M": "comments",
}

REPLACEMENTS = [
    ("Это упражнения было", "Это упражнение было"),
    ("превезено", "привезено"),
    ("получаться", "получатся"),
    ("В течении", "В течение"),
    ("наружу это", "наружу, это"),
    ("обойти,а", "обойти, а"),
    ("сражение", "сражения"),
    ("падишаха это", "падишаха - это"),
    ("государство ввязалась", "государство ввязалось"),
    ("Если родиться мальчик", "Если родится мальчик"),
    ("Если родиться девочка", "Если родится девочка"),
    ("становятся ведущим", "становится ведущим"),
    ("бережно относится", "бережно относиться"),
    ("обе команда", "обе команды"),
    ("Внешний круг это", "Внешний круг - это"),
    ("Внутренний круг это", "Внутренний круг - это"),
    ("Это игра была", "Эта игра была"),
    ("учасников", "участников"),
    ("того значения, на котором остановился в предыдущий раз", "того значения, на котором он остановился в предыдущий раз"),
    ("идет игра", "идёт игра"),
    ("по другому", "по-другому"),
    ("не нем", "на нём"),
    ("именуется королем", "именуется королём"),
    ("тот кто", "тот, кто"),
    ("садишься", "садитесь"),
    ("садиться (выбывает)", "садится (выбывает)"),
    ("ошиабется", "ошибается"),
    ("рандомно", "хаотично"),
    ("некому пространству", "некому пространству"),
    ("достаточной длинны", "достаточной длины"),
    ("челнам своей команды", "членам своей команды"),
    ("по разному", "по-разному"),
    ("нравиться", "нравится"),
    ("становиться суперменом", "становится суперменом"),
    ("не стали суперменами", "не стали суперменами"),
    ("садиться", "садится"),
    ("а отведенное время", "за отведённое время"),
    ("подумать на башню", "подуть на башню"),
    ("развится", "развиться"),
    ("марярный", "малярный"),
    ("что бы", "чтобы"),
    ("выехавщие", "выехавшие"),
    ("за пределя", "за пределы"),
    ("строиться в линию", "строится в линию"),
    ("хочеться", "хочется"),
    ("стоить в линии", "стоять в линии"),
    ("выстраиваться в очередь", "выстраиваются в очередь"),
    ("свзяны", "связаны"),
    ("договариватся", "договариваться"),
    ("учасника", "участника"),
    ("касатся её", "касаться её"),
    ("сходят по помещению", "ходят по помещению"),
    ("напомнить", "напомнить"),
    ("который делают", "которые делают"),
    ("акварельные открытки(рисуем", "акварельные открытки (рисуем"),
    ("осенний коллаж(ассоциации", "осенний коллаж (ассоциации"),
    ("бусины с воспоминаниями(каждая", "бусины с воспоминаниями (каждая"),
    ("учасников", "участников"),
    ("монстра. Рассказывают", "монстра. Ведущий рассказывает"),
    ("моснсер", "монстр"),
    ("уводик", "уводит"),
    ("на мг", "на МГ"),
    ("запомнить имена участников", "Запомнить имена участников"),
]

REGEX_REPLACEMENTS = [
    (r"\b(не ограничено)\b", "Не ограничено"),
    (r"\b(любой)\b", "Любой"),
    (r" {2,}", " "),
    (r" *\n", "\n"),
    (r"\n{3,}", "\n\n"),
]


def normalize_text(value: str) -> str:
    text = (value or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return ""

    for old, new in REPLACEMENTS:
        text = text.replace(old, new)

    for pattern, replacement in REGEX_REPLACEMENTS:
        text = re.sub(pattern, replacement, text, flags=re.MULTILINE)

    text = text.replace('"коллективный рисунок"', "«коллективный рисунок»")
    text = text.replace('"сижу"', "«сижу»")
    text = text.strip()
    return text


def normalize_location(value: str) -> str:
    text = normalize_text(value)
    if text in DEFAULT_LOCATIONS:
        return text
    return "Не имеет значения" if not text else text


def restore_rules_column(rules: str, comments: str, rules_link: str) -> str:
    cleaned = rules
    if comments:
        suffix = f"\n\nКомментарий:\n{comments}"
        if cleaned.endswith(suffix):
            cleaned = cleaned[: -len(suffix)]
    if rules_link:
        suffix = f"\n\nСсылка на правила:\n{rules_link}"
        if cleaned.endswith(suffix):
            cleaned = cleaned[: -len(suffix)]
    return cleaned.strip()


def workbook_rows():
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[wb.sheetnames[0]]

    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for column, field in TEXT_COLUMNS.items():
            row_data[field] = normalize_text(ws[f"{column}{row_idx}"].value or "")

        if not row_data["title"]:
            continue

        row_data["rules"] = restore_rules_column(
            row_data["rules"],
            row_data["comments"],
            row_data["rules_link"],
        )

        yield row_idx, {
            "title": row_data["title"],
            "game_type": row_data["game_type"],
            "goal": row_data["goal"],
            "participants": row_data["participants"] or "Не ограничено",
            "age_category": row_data["age_category"] or "Любой",
            "duration": row_data["duration"],
            "location": normalize_location(row_data["location"]),
            "equipment": row_data["equipment"],
            "rules": row_data["rules"],
            "rules_link": row_data["rules_link"],
            "comments": row_data["comments"],
        }


def update_workbook() -> int:
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[wb.sheetnames[0]]
    changed = 0

    for row_idx, normalized in workbook_rows():
        updates = {
            "C": normalized["title"],
            "D": normalized["game_type"],
            "E": normalized["goal"],
            "F": normalized["participants"],
            "G": normalized["age_category"],
            "H": normalized["duration"],
            "I": normalized["location"],
            "J": normalized["equipment"],
            "K": normalized["rules"],
            "L": normalized["rules_link"],
            "M": normalized["comments"],
        }

        for column, new_value in updates.items():
            cell = ws[f"{column}{row_idx}"]
            current = "" if cell.value is None else str(cell.value).replace("\r\n", "\n").replace("\r", "\n").strip()
            if current != new_value:
                cell.value = new_value
                changed += 1

    wb.save(WORKBOOK_PATH)
    return changed


def import_into_db() -> int:
    conn = sqlite3.connect(DATABASE_PATH)
    cur = conn.cursor()

    cur.execute("DELETE FROM games")
    cur.execute("DELETE FROM game_types")
    cur.execute("DELETE FROM age_categories")

    imported = 0
    game_types: set[str] = set(DEFAULT_GAME_TYPES)
    age_categories: set[str] = set(DEFAULT_AGE_OPTIONS)

    for _row_idx, row in workbook_rows():
        merged_rules = row["rules"]
        if row["comments"]:
            merged_rules = f"{merged_rules}\n\nКомментарий:\n{row['comments']}".strip()
        if row["rules_link"]:
            merged_rules = f"{merged_rules}\n\nСсылка на правила:\n{row['rules_link']}".strip()

        cur.execute(
            """
            INSERT INTO games (
                title, game_type, goal, participants, age_category,
                duration, location, equipment, rules, files_json
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, '[]')
            """,
            (
                row["title"],
                row["game_type"],
                row["goal"],
                row["participants"],
                row["age_category"],
                row["duration"],
                row["location"],
                row["equipment"],
                merged_rules,
            ),
        )
        imported += 1
        if row["game_type"]:
            for game_type in [item.strip() for item in row["game_type"].split(",") if item.strip()]:
                game_types.add(game_type)
        if row["age_category"]:
            age_categories.add(row["age_category"])

    for game_type in sorted(game_types):
        cur.execute("INSERT OR IGNORE INTO game_types (name) VALUES (?)", (game_type,))

    for age_category in sorted(age_categories):
        cur.execute("INSERT OR IGNORE INTO age_categories (name) VALUES (?)", (age_category,))

    conn.commit()
    conn.close()
    return imported


if __name__ == "__main__":
    changed_cells = update_workbook()
    imported_rows = import_into_db()
    print(f"Updated workbook cells: {changed_cells}")
    print(f"Imported rows: {imported_rows}")
